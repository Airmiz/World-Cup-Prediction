"""Build predictions.xlsx from model artifacts."""
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
HDR_FILL = PatternFill("solid", start_color="1F4E79")
ALT_FILL = PatternFill("solid", start_color="EAF1F8")
PCT = "0.0%"
thin = Border(bottom=Side(style="thin", color="D9D9D9"))

mp = pd.read_csv("output/match_predictions.csv")
tp = pd.read_csv("output/team_probabilities.csv")
bt = json.load(open("output/backtest.json"))
ss = json.load(open("output/sim_summary.json"))

wb = Workbook()

def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)

def write_df(ws, df, pct_cols=(), num_cols=(), widths=None, start_row=1):
    ws.append(list(df.columns)) if start_row == 1 else None
    for _, r in df.iterrows():
        ws.append(list(r.values))
    style_header(ws, len(df.columns), row=start_row)
    for i, _ in enumerate(df.columns, 1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = (widths or {}).get(i, 14)
    for ri in range(start_row + 1, start_row + 1 + len(df)):
        for ci in range(1, len(df.columns) + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.font = Font(name=ARIAL, size=10)
            cell.border = thin
            if (ri - start_row) % 2 == 0:
                cell.fill = ALT_FILL
            if ci in pct_cols:
                cell.number_format = PCT
            elif ci in num_cols:
                cell.number_format = "0.00"

# ---------------- Overview ----------------
ws = wb.active; ws.title = "Overview"
ws["A1"] = "FIFA World Cup 2026 — Statistical Forecast"
ws["A1"].font = Font(name=ARIAL, bold=True, size=16, color="1F4E79")
lines = [
    ("Model", "Time-decayed Dixon-Coles bivariate Poisson blended 50/50 (log-linear) with World Football Elo goal rates"),
    ("Training data", "49,477 international matches, 1872 – 10 Jun 2026 (martj42/international_results)"),
    ("Simulations", f"{ss['n_sims']:,} full-tournament Monte Carlo runs (seed {ss['seed']})"),
    ("Knockout rules", "Official bracket incl. FIFA Annex C third-place allocation (495 combinations), extra time + penalties"),
    ("Home advantage", f"+{ss['dc_home_adv']:.2f} log-goals for host nations playing in their own country"),
    ("Validation", "Out-of-sample on WC 2014/2018/2022 group stages; hyperparameters tuned on 2014+2018 only, 2022 held out"),
    ("Held-out 2022 log loss", f"{bt['2022']['blend']['log_loss']:.3f} (uniform baseline {bt['2022']['uniform']['log_loss']:.3f}; Elo-only {bt['2022']['elo_only']['log_loss']:.3f})"),
    ("Generated", "11 June 2026, pre-tournament"),
]
for i, (k, v) in enumerate(lines, start=3):
    ws[f"A{i}"] = k; ws[f"B{i}"] = v
    ws[f"A{i}"].font = Font(name=ARIAL, bold=True, size=10)
    ws[f"B{i}"].font = Font(name=ARIAL, size=10)
ws.column_dimensions["A"].width = 24; ws.column_dimensions["B"].width = 110

ws["A12"] = "Title favourites"
ws["A12"].font = Font(name=ARIAL, bold=True, size=12, color="1F4E79")
ws.append([]); ws.append(["Rank", "Team", "P(champion)"])
for i, r in enumerate(ss["champion_top10"], 1):
    ws.append([i, r["team"], r["p_champion"]])
style_header(ws, 3, row=14)
for ri in range(15, 25):
    ws.cell(row=ri, column=3).number_format = PCT
    for ci in (1, 2, 3):
        ws.cell(row=ri, column=ci).font = Font(name=ARIAL, size=10)

# ---------------- Match predictions ----------------
ws = wb.create_sheet("Match Predictions")
m = mp.rename(columns={
    "date": "Date", "group": "Group", "home_team": "Home", "away_team": "Away",
    "city": "City", "country": "Country", "p_home_win": "P(Home win)",
    "p_draw": "P(Draw)", "p_away_win": "P(Away win)",
    "exp_goals_home": "xG Home", "exp_goals_away": "xG Away",
    "most_likely_scores": "Most likely scores"})
write_df(ws, m, pct_cols=(7, 8, 9), num_cols=(10, 11),
         widths={1: 11, 2: 7, 3: 18, 4: 18, 5: 14, 6: 13, 12: 34})

# ---------------- Team probabilities ----------------
ws = wb.create_sheet("Team Probabilities")
t = tp.rename(columns={
    "team": "Team", "group": "Group", "elo": "Elo", "dc_attack": "Attack",
    "dc_defence": "Defence", "exp_group_points": "Exp. pts",
    "p_win_group": "P(win group)", "p_group_2nd": "P(2nd)", "p_group_3rd": "P(3rd)",
    "p_group_4th": "P(4th)", "p_reach_r32": "P(R32)", "p_reach_r16": "P(R16)",
    "p_reach_qf": "P(QF)", "p_reach_sf": "P(SF)", "p_reach_final": "P(Final)",
    "p_champion": "P(Champion)"})
t["Elo"] = t["Elo"].round(0)
write_df(ws, t, pct_cols=tuple(range(7, 17)), num_cols=(4, 5, 6),
         widths={1: 18, 2: 7, 3: 8})

# ---------------- Groups ----------------
ws = wb.create_sheet("Groups")
row = 1
for g in "ABCDEFGHIJKL":
    sub = tp[tp["group"] == g].sort_values("exp_group_points", ascending=False)
    ws.cell(row=row, column=1, value=f"Group {g}").font = Font(name=ARIAL, bold=True, size=12, color="1F4E79")
    row += 1
    hdr = ["Team", "Elo", "Exp. pts", "P(1st)", "P(2nd)", "P(3rd)", "P(4th)", "P(advance)"]
    for ci, h in enumerate(hdr, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = Font(name=ARIAL, bold=True, color="FFFFFF", size=10); c.fill = HDR_FILL
    row += 1
    for _, r in sub.iterrows():
        vals = [r["team"], round(r["elo"]), r["exp_group_points"], r["p_win_group"],
                r["p_group_2nd"], r["p_group_3rd"], r["p_group_4th"], r["p_reach_r32"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = Font(name=ARIAL, size=10)
            if ci == 3: c.number_format = "0.00"
            if ci >= 4: c.number_format = PCT
        row += 1
    row += 1
for i, wdt in enumerate([18, 8, 9, 9, 9, 9, 9, 11], 1):
    ws.column_dimensions[get_column_letter(i)].width = wdt

# ---------------- Validation ----------------
ws = wb.create_sheet("Validation")
ws["A1"] = "Out-of-sample backtest — World Cup group stages"
ws["A1"].font = Font(name=ARIAL, bold=True, size=13, color="1F4E79")
ws["A2"] = "Hyperparameters tuned on 2014+2018; 2022 fully held out. Lower is better. Uniform = always (1/3, 1/3, 1/3)."
ws["A2"].font = Font(name=ARIAL, italic=True, size=9)
ws.append([]); ws.append(["Tournament", "Model", "Log loss", "Brier", "Matches"])
r0 = ws.max_row
for y in ("2014", "2018", "2022"):
    for k, label in [("blend", "Blend (final)"), ("dc_only", "Dixon-Coles only"),
                     ("elo_only", "Elo only"), ("uniform", "Uniform baseline")]:
        d = bt[y][k]
        ws.append([y, label, round(d["log_loss"], 4), round(d["brier"], 4), d["n"]])
style_header(ws, 5, row=r0)
for ri in range(r0 + 1, ws.max_row + 1):
    for ci in range(1, 6):
        c = ws.cell(row=ri, column=ci); c.font = Font(name=ARIAL, size=10)
        if ws.cell(row=ri, column=2).value == "Blend (final)":
            c.font = Font(name=ARIAL, size=10, bold=True)
for i, wdt in enumerate([12, 20, 10, 10, 10], 1):
    ws.column_dimensions[get_column_letter(i)].width = wdt
ws.append([])
ws.append(["Tuned hyperparameters:", f"time decay xi = {bt['best']['xi']}/yr, friendly weight = {bt['best']['friendly_w']}, blend weight = {bt['best']['blend_w']}"])
ws.cell(row=ws.max_row, column=1).font = Font(name=ARIAL, bold=True, size=10)
ws.cell(row=ws.max_row, column=2).font = Font(name=ARIAL, size=10)

wb.save("output/WC2026_predictions.xlsx")
print("saved output/WC2026_predictions.xlsx")
